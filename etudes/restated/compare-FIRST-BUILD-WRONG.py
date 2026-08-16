#!/usr/bin/env python3
"""Étude instrument: does an already-published figure about a fixed past month
change in a later edition of the same register?

Reads the archived, date-stamped editions of ICE's FY2025 detention statistics
workbook and compares the sheet ' ICLOS and Detainees', whose every cell is keyed
to a calendar month that is already over. Any difference between two editions on
the same key is the register giving a different answer to the same question about
the same past month.

Deterministic: same inputs, same output. No network.
"""
import json, os, sys, glob, hashlib
import openpyxl

SHEET = " ICLOS and Detainees"
CAPS = "caps"


def ffill(row):
    """Merged header cells read as None; carry the last value forward."""
    out, last = [], None
    for c in row:
        if c is not None and str(c).strip() != "":
            last = str(c).strip()
        out.append(last)
    return out


def parse(path):
    """-> {(block, population, year, month, point): value}"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        return None
    ws = wb[SHEET]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    vals = {}
    block = None
    years = months = points = None
    for r in rows:
        first = (str(r[0]).strip() if r[0] is not None else "")
        # a header stack starts at a cell literally labelled "Population"
        if first.lower().startswith("population"):
            years = ffill(r)
            continue
        if years is not None and months is None:
            months = ffill(r)
            continue
        if months is not None and points is None:
            points = ffill(r)
            continue
        if first == "" or r[0] is None:
            continue
        # a text-only row with no numbers below a completed header = block label
        numeric = [(i, c) for i, c in enumerate(r) if isinstance(c, (int, float))]
        if not numeric:
            if years is None:
                block = first
            else:
                block = first  # sub-block label inside a matrix (e.g. "Adult Facility")
            continue
        if years is None or months is None or points is None:
            continue
        for i, c in numeric:
            y, m, p = (years[i] if i < len(years) else None), \
                      (months[i] if i < len(months) else None), \
                      (points[i] if i < len(points) else None)
            if not (y and m and p):
                continue
            if not y.isdigit():
                continue
            vals[(block or "", first, y, m, p)] = float(c)
    return vals


def edition_date(name):
    # FY25_detentionStats03052025.xlsx -> 2025-03-05
    d = name.replace("FY25_detentionStats", "").replace(".xlsx", "")
    if len(d) != 8 or not d.isdigit():
        return None
    return f"{d[4:]}-{d[0:2]}-{d[2:4]}"


def main():
    man = {m["file"]: m for m in json.load(open(f"{CAPS}/manifest.json"))}
    eds = []
    for f in sorted(glob.glob(f"{CAPS}/FY25_detentionStats*.xlsx")):
        name = os.path.basename(f)
        d = edition_date(name)
        if not d:
            continue
        v = parse(f)
        if v is None:
            print(f"  (no sheet {SHEET!r} in {name} — skipped)", file=sys.stderr)
            continue
        eds.append({"file": name, "date": d, "vals": v,
                    "sha256": man.get(name, {}).get("sha256")})
    eds.sort(key=lambda e: e["date"])
    print(f"editions parsed: {len(eds)}  ({eds[0]['date']} .. {eds[-1]['date']})")
    print(f"keys in first edition: {len(eds[0]['vals'])}, in last: {len(eds[-1]['vals'])}")

    # For every key, walk the editions in order and record every value change.
    first_seen, changes = {}, []
    for e in eds:
        for k, v in e["vals"].items():
            if k not in first_seen:
                first_seen[k] = (e["date"], v)
                continue
            prev_date, prev_v = first_seen[k]
            if v != prev_v:
                changes.append({"key": list(k), "from": prev_v, "to": v,
                                "first_published": prev_date, "changed_in": e["date"]})
            first_seen[k] = (e["date"], v)

    tracked = len(first_seen)
    keys_changed = {tuple(c["key"]) for c in changes}
    print(f"\nkeys tracked across editions: {tracked}")
    print(f"keys whose published value changed at least once: {len(keys_changed)}"
          f"  ({100*len(keys_changed)/tracked:.1f} %)")
    print(f"total change events: {len(changes)}")

    # How far back does a rewritten month lie?
    out = {"editions": [{"file": e["file"], "date": e["date"], "sha256": e["sha256"],
                         "keys": len(e["vals"])} for e in eds],
           "keys_tracked": tracked, "keys_changed": len(keys_changed),
           "change_events": len(changes), "changes": changes}
    json.dump(out, open("etude-result.json", "w"), indent=1)

    print("\nfirst 25 change events:")
    for c in changes[:25]:
        b, pop, y, m, p = c["key"]
        print(f"  {y} {m} {p:4s} | {b[:22]:22s} | {pop[:26]:26s} | "
              f"{c['from']:>12} -> {c['to']:<12} | published {c['first_published']}"
              f" changed {c['changed_in']}")


if __name__ == "__main__":
    main()
