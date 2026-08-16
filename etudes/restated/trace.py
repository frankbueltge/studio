#!/usr/bin/env python3
"""Étude instrument — trace one row of the register across every archived edition.

The question this answers: when the register republishes a figure about a month
that is already over, does the answer stay the same?

Usage:  python3 trace.py [month-point ...]        e.g.  2024-Aug-end 2024-May-end

Offline. Reads the workbooks named in corpus-manifest.json from ./caps/, which
`fetch.py` downloads from the Internet Archive. The corpus is NOT committed
(19 MB of a third party's files); the manifest carries each file's sha256 and its
archive timestamp so anyone can rebuild the identical corpus and re-run this.
"""
import glob, json, os, sys
import openpyxl

SHEET = " ICLOS and Detainees"
ROW = "Single Adults with a Positive Fear Determination Still in Custody"
DEFAULT_KEYS = ["2024-Aug-end", "2024-May-end", "2023-Oct-mid"]


def ffill(row):
    out, last = [], None
    for c in row:
        if c is not None and str(c).strip():
            last = str(c).strip()
        out.append(last)
    return out


def series(path, label=ROW):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        return None
    rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    yr, mo, pt = ffill(rows[3]), ffill(rows[4]), ffill(rows[5])
    for r in rows[:12]:
        if r and r[0] == label:
            return {f"{yr[i]}-{mo[i][:3]}-{pt[i].strip()}": c
                    for i, c in enumerate(r)
                    if isinstance(c, (int, float)) and yr[i] and mo[i] and pt[i]}
    return {}


def edition_date(name):
    s = name.replace("FY25_detentionStats", "").replace(".xlsx", "")
    return f"{s[4:]}-{s[0:2]}-{s[2:4]}"


def main():
    keys = sys.argv[1:] or DEFAULT_KEYS
    files = sorted(glob.glob("caps/FY25_detentionStats*.xlsx"),
                   key=lambda p: edition_date(os.path.basename(p)))
    if not files:
        sys.exit("no corpus in ./caps — run `python3 fetch.py` first")
    print(f"row: {ROW}")
    print(f"unit: average days in custody\n")
    print(f"{'edition published':18s}" + "".join(f"{k:>15s}" for k in keys))
    for f in files:
        s = series(f)
        if s is None:
            continue
        d = edition_date(os.path.basename(f))
        print(f"{d:18s}" + "".join(
            (f"{s[k]:15.2f}" if k in s else f"{'—':>15s}") for k in keys))


if __name__ == "__main__":
    main()
