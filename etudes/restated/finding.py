#!/usr/bin/env python3
"""Étude instrument — the finding, stated as the material carries it.

Written to discharge the Kritiker's condition 4 at the session-98 concept gate:
the finding must be recomputable by a second hand from committed data, and it must
say what the material says rather than what the first draft dramatised.

Definition, stated because the first attempt hid it: a MATERIAL RESTATEMENT is a
key (row · year · month · mid-or-end) whose FIRST non-zero published value and
LAST non-zero published value differ by more than 5 %. The zero editions are
excluded by that definition, deliberately — see ETUDE-01.md, "the zero editions".

Offline. Reads ./caps/, which fetch.py rebuilds from corpus-manifest.json.
"""
import collections, glob, os
import openpyxl

SHEET = " ICLOS and Detainees"
MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
THRESHOLD = 0.05


def ffill(row):
    out, last = [], None
    for c in row:
        if c is not None and str(c).strip():
            last = str(c).strip()
        out.append(last)
    return out


def edition_date(name):
    s = name.replace("FY25_detentionStats", "").replace(".xlsx", "")
    return f"{s[4:]}-{s[0:2]}-{s[2:4]}"


def history():
    """-> {(row, year, mon, point): [(edition_date, value), ...]}"""
    hist = collections.defaultdict(list)
    files = sorted(glob.glob("caps/FY25_detentionStats*.xlsx"),
                   key=lambda p: edition_date(os.path.basename(p)))
    for f in files:
        ed = edition_date(os.path.basename(f))
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        if SHEET not in wb.sheetnames:
            continue
        rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
        yr, mo, pt = ffill(rows[3]), ffill(rows[4]), ffill(rows[5])
        for r in rows[6:11]:                       # the ICLOS block's data rows
            lab = r[0]
            if not isinstance(lab, str) or not lab.strip():
                continue
            for i, c in enumerate(r):
                if (isinstance(c, (int, float)) and i < len(yr)
                        and yr[i] and yr[i].isdigit() and mo[i] and pt[i]):
                    hist[(lab.strip(), yr[i], mo[i][:3], pt[i].strip())].append((ed, c))
    return hist, len(files)


def main():
    hist, n_ed = history()
    material, untouched_fy23 = [], []
    for k, v in hist.items():
        nz = [(e, x) for e, x in v if x != 0]
        if k[1] == "2023" and MON.index(k[2]) < 9:
            untouched_fy23.append((k, {round(x, 6) for _, x in v}))
        if len(nz) < 2:
            continue
        first, last = nz[0][1], nz[-1][1]
        if first and abs(last - first) / abs(first) > THRESHOLD:
            material.append((k, first, last, nz[0][0], nz[-1][0]))

    rows = collections.Counter(k[0] for k, *_ in material)
    up = sum(1 for _, f, l, *_ in material if l > f)
    down = len(material) - up

    print(f"editions read: {n_ed}    keys with any published value: {len(hist)}")
    print(f"\nMATERIAL RESTATEMENTS (>{THRESHOLD:.0%}, first vs last non-zero): "
          f"{len(material)}")
    print(f"  rows they occur in: {len(rows)}")
    for r, n in rows.items():
        print(f"    {n:3d}  {r}")
    print(f"  direction: {up} revised UP, {down} revised DOWN")
    byyear = collections.Counter(k[1] for k, *_ in material)
    print(f"  by year described: {dict(sorted(byyear.items()))}")

    print(f"\nJanuary–September 2023 (the months OUTSIDE the rewritten window):")
    multi = [k for k, vals in untouched_fy23 if len(vals) > 1]
    print(f"  points: {len(untouched_fy23)}   points that ever took more than one "
          f"value across all {n_ed} editions: {len(multi)}")

    print("\nevery material restatement:")
    for k, f_, l_, e1, e2 in sorted(material, key=lambda x: (x[0][1], MON.index(x[0][2]))):
        print(f"  {k[1]}-{k[2]}-{k[3]:4s} {f_:8.2f} -> {l_:8.2f}  "
              f"({100*(l_-f_)/f_:+6.1f} %)   {e1} -> {e2}")


if __name__ == "__main__":
    main()
