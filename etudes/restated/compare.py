#!/usr/bin/env python3
"""Etude instrument, second build.

First build reported 38.1 % of keys changed. It was wrong: its keys collided
across the sheet's stacked blocks, so a value from one block was compared against
a different block's value in the next edition, and whole runs of "-> 0.0" appeared
where the raw cells are in fact identical. That failure is kept in the record.

This build refuses to compare anything it cannot prove unique:
  * header rows (year / month / mid-end) are re-read for EVERY block,
  * each block gets an ordinal, so two blocks with the same row label cannot merge,
  * any key occurring more than once inside a single edition is dropped as
    AMBIGUOUS from every edition, not silently kept.

Deterministic, offline. Same inputs, same output.
"""
import json, os, sys, glob
from collections import Counter
import openpyxl

SHEET = " ICLOS and Detainees"
CAPS = "caps"
MONTHS = {"january", "february", "march", "april", "may", "june", "july",
          "august", "september", "october", "november", "december"}


def ffill(row):
    out, last = [], None
    for c in row:
        if c is not None and str(c).strip() != "":
            last = str(c).strip()
        out.append(last)
    return out


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        return None, None
    ws = wb[SHEET]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    vals = {}
    block_ord = 0
    years = months = points = None
    need = 0          # 1 = expect month row, 2 = expect mid/end row
    for r in rows:
        first = (str(r[0]).strip() if r[0] is not None else "")
        low = first.lower()

        if low.startswith("population"):
            block_ord += 1
            years, months, points = ffill(r), None, None
            need = 1
            continue
        if need == 1:
            # month row: recognisable by containing month names
            if sum(1 for c in r if c and str(c).strip().lower() in MONTHS) >= 2:
                months = ffill(r); need = 2
            continue
        if need == 2:
            points = ffill(r); need = 0
            continue

        numeric = [(i, c) for i, c in enumerate(r) if isinstance(c, (int, float))]
        if not numeric or years is None or months is None or points is None:
            continue
        for i, c in numeric:
            if i >= len(years) or i >= len(months) or i >= len(points):
                continue
            y, m, p = years[i], months[i], points[i]
            if not (y and m and p) or not y.isdigit():
                continue
            if m.lower() not in MONTHS:
                continue
            key = (block_ord, first, y, m, p)
            vals.setdefault(key, []).append(float(c))

    dupes = {k for k, v in vals.items() if len(v) > 1}
    clean = {k: v[0] for k, v in vals.items() if len(v) == 1}
    return clean, dupes


def edition_date(name):
    d = name.replace("FY25_detentionStats", "").replace(".xlsx", "")
    return f"{d[4:]}-{d[0:2]}-{d[2:4]}" if len(d) == 8 and d.isdigit() else None


def main():
    man = {m["file"]: m for m in json.load(open(f"{CAPS}/manifest.json"))}
    eds, ambiguous = [], set()
    for f in sorted(glob.glob(f"{CAPS}/FY25_detentionStats*.xlsx")):
        name = os.path.basename(f)
        d = edition_date(name)
        if not d:
            continue
        v, dup = parse(f)
        if v is None:
            print(f"  skipped (no sheet): {name}", file=sys.stderr); continue
        ambiguous |= dup
        eds.append({"file": name, "date": d, "vals": v,
                    "sha256": man.get(name, {}).get("sha256")})
    eds.sort(key=lambda e: e["date"])
    for e in eds:
        for k in ambiguous:
            e["vals"].pop(k, None)

    print(f"editions: {len(eds)}  {eds[0]['date']} .. {eds[-1]['date']}")
    print(f"keys dropped as ambiguous (collided inside some edition): {len(ambiguous)}")

    last_seen, changes = {}, []
    for e in eds:
        for k, v in e["vals"].items():
            if k not in last_seen:
                last_seen[k] = (e["date"], v); continue
            pd_, pv = last_seen[k]
            if v != pv:
                changes.append({"block": k[0], "row": k[1], "year": k[2], "month": k[3],
                                "point": k[4], "from": pv, "to": v,
                                "previous_edition": pd_, "changed_in": e["date"]})
            last_seen[k] = (e["date"], v)

    tracked = len(last_seen)
    changed = {(c["block"], c["row"], c["year"], c["month"], c["point"]) for c in changes}
    print(f"keys tracked: {tracked}")
    print(f"keys whose already-published value later changed: {len(changed)}"
          f"  ({100*len(changed)/tracked:.1f} %)")
    print(f"change events: {len(changes)}")
    if changes:
        yrs = Counter(c["year"] for c in changes)
        print("change events by the YEAR the figure describes:", dict(sorted(yrs.items())))
        print("\nlargest 20 by absolute size:")
        for c in sorted(changes, key=lambda c: -abs(c["to"] - c["from"]))[:20]:
            print(f"  {c['year']} {c['month']:9s} {c['point']:4s} | blk{c['block']} "
                  f"{c['row'][:30]:30s} | {c['from']:>10.2f} -> {c['to']:<10.2f} "
                  f"| prev {c['previous_edition']} chg {c['changed_in']}")

    json.dump({"editions": [{k: e[k] for k in ("file", "date", "sha256")} | {"keys": len(e["vals"])}
                            for e in eds],
               "ambiguous_keys_dropped": len(ambiguous),
               "keys_tracked": tracked, "keys_changed": len(changed),
               "change_events": len(changes), "changes": changes},
              open("etude-result.json", "w"), indent=1)


if __name__ == "__main__":
    main()
